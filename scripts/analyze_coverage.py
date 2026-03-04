#!/usr/bin/env python3
import os
import re
import sys
from pathlib import Path

def detect_language(repo_path):
    file_counts = {
        'cpp': 0,
        'python': 0,
        'javascript': 0,
        'java': 0,
        'go': 0,
    }
    
    for ext in ['*.cpp', '*.hpp', '*.h', '*.c']:
        file_counts['cpp'] += len(list(repo_path.rglob(ext)))
    for ext in ['*.py']:
        file_counts['python'] += len(list(repo_path.rglob(ext)))
    for ext in ['*.js', '*.ts']:
        file_counts['javascript'] += len(list(repo_path.rglob(ext)))
    for ext in ['*.java']:
        file_counts['java'] += len(list(repo_path.rglob(ext)))
    for ext in ['*.go']:
        file_counts['go'] += len(list(repo_path.rglob(ext)))
    
    return max(file_counts, key=file_counts.get)

def get_source_dirs(repo_path, language):
    common_dirs = [
        repo_path / "include",
        repo_path / "src",
        repo_path / "lib",
        repo_path / "api",
        repo_path / "public",
        repo_path / "source",
    ]
    
    language_specific = {
        'cpp': [repo_path / "include", repo_path / "src"],
        'python': [repo_path / "src", repo_path / "lib"],
        'javascript': [repo_path / "src", repo_path / "lib"],
        'java': [repo_path / "src", repo_path / "main"],
        'go': [repo_path / "pkg", repo_path / "src"],
    }
    
    dirs = language_specific.get(language, common_dirs)
    return [d for d in dirs if d.exists()]

def get_test_dirs(repo_path):
    test_dirs = [
        repo_path / "tests",
        repo_path / "test",
        repo_path / "unittest",
        repo_path / "spec",
        repo_path / "__tests__",
        repo_path / "test_cases",
    ]
    return [d for d in test_dirs if d.exists()]

def get_source_extensions(language):
    extensions = {
        'cpp': ['*.h', '*.hpp', '*.c', '*.cpp'],
        'python': ['*.py'],
        'javascript': ['*.js', '*.ts'],
        'java': ['*.java'],
        'go': ['*.go'],
    }
    return extensions.get(language, ['*.h', '*.hpp', '*.py', '*.js', '*.ts'])

def get_test_extensions(language):
    extensions = {
        'cpp': ['*test*.cpp', '*test*.c', '*_test.cpp', '*_test.c'],
        'python': ['test*.py', '*_test.py'],
        'javascript': ['*.test.js', '*.test.ts', '*.spec.js'],
        'java': ['*Test.java', '*Tests.java'],
        'go': ['*_test.go'],
    }
    return extensions.get(language, ['*test*'])

def get_api_patterns(language):
    patterns = {
        'cpp': [
            r'PUBLIC_API\s+(\w+)\s*\([^)]*\)',
            r'EXPORT_API\s+(\w+)\s*\([^)]*\)',
            r'DLLEXPORT\s+(\w+)\s*\([^)]*\)',
            r'extern\s+"C"\s*\{?\s*(\w+)\s*\([^)]*\)',
            r'__declspec\(dllexport\)\s+(\w+)\s*\([^)]*\)',
            r'ACLSHMEM_HOST_API\s+(\w+)\s*\([^)]*\)',
            r'ACLSHMEM_DEVICE\s+(\w+)\s*\([^)]*\)',
            r'ACLSHMEM_API\s+(\w+)\s*\([^)]*\)',
        ],
        'python': [
            r'@public\s+def\s+(\w+)\s*\(',
            r'@api\s+def\s+(\w+)\s*\(',
            r'@export\s+def\s+(\w+)\s*\(',
            r'def\s+(\w+)\s*\(',  # All functions
        ],
        'javascript': [
            r'export\s+function\s+(\w+)\s*\(',
            r'export\s+const\s+(\w+)\s*=',
            r'module\.exports\.\s*(\w+)\s*=',
        ],
        'java': [
            r'public\s+(?:\w+)\s+(\w+)\s*\([^)]*\)',
            r'@Api\s+public\s+(\w+)\s*\([^)]*\)',
        ],
        'go': [
            r'func\s+([A-Z]\w+)\s*\(',  # Exported functions start with capital
        ],
    }
    return patterns.get(language, [
        r'PUBLIC_API\s+(\w+)\s*\([^)]*\)',
        r'EXPORT_API\s+(\w+)\s*\([^)]*\)',
        r'DLLEXPORT\s+(\w+)\s*\([^)]*\)',
        r'extern\s+"C"\s*\{?\s*(\w+)\s*\([^)]*\)',
        r'ACLSHMEM_HOST_API\s+(\w+)\s*\([^)]*\)',
        r'ACLSHMEM_DEVICE\s+(\w+)\s*\([^)]*\)',
    ])

def get_category_from_path(file_path):
    path_str = str(file_path).lower()
    
    category_map = {
        'init': 'Initialization',
        'setup': 'Initialization',
        'config': 'Configuration',
        'mem': 'Memory',
        'alloc': 'Memory',
        'heap': 'Memory',
        'io': 'I/O',
        'file': 'I/O',
        'net': 'Communication',
        'comm': 'Communication',
        'sync': 'Synchronization',
        'thread': 'Concurrency',
        'util': 'Utilities',
        'helper': 'Utilities',
        'tool': 'Utilities',
        'math': 'Math',
        'calc': 'Calculation',
        'data': 'Data',
        'model': 'Data',
        'db': 'Database',
        'storage': 'Storage',
        'auth': 'Authentication',
        'login': 'Authentication',
        'user': 'User',
        'account': 'Account',
        'api': 'API',
        'service': 'Service',
        'controller': 'Controller',
        'handler': 'Handler',
        'host': 'Host',
        'device': 'Device',
        'rma': 'RMA',
        'amo': 'Atomic',
        'cc': 'Collective',
        'p2p': 'P2P',
        'team': 'Team',
        'so': 'Signal',
        'mo': 'Memory',
    }
    
    for keyword, category in category_map.items():
        if keyword in path_str:
            return category
    
    return 'Other'

def find_source_files(repo_path, language):
    source_dirs = get_source_dirs(repo_path, language)
    extensions = get_source_extensions(language)
    
    source_files = []
    for source_dir in source_dirs:
        for ext in extensions:
            for file in source_dir.rglob(ext):
                source_files.append(file)
    
    if not source_files:
        for ext in extensions:
            for file in repo_path.rglob(ext):
                source_files.append(file)
    
    return source_files

def find_test_files(repo_path, language):
    test_dirs = get_test_dirs(repo_path)
    extensions = get_test_extensions(language)
    
    test_files = []
    for test_dir in test_dirs:
        for ext in extensions:
            for file in test_dir.rglob(ext):
                test_files.append(file)
    
    return test_files

def extract_apis_from_source(source_file, repo_path, language):
    apis = []
    
    try:
        content = source_file.read_text(encoding='utf-8', errors='ignore')
    except Exception:
        return apis
    
    category = get_category_from_path(source_file)
    api_patterns = get_api_patterns(language)
    
    if language == 'cpp':
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('/*'):
                continue
            
            for pattern in api_patterns:
                matches = re.findall(pattern, line)
                for match in matches:
                    api_name = match
                    if api_name and not api_name.startswith('_'):
                        apis.append({
                            'name': api_name,
                            'source': source_file.name,
                            'category': category,
                            'full_path': str(source_file.relative_to(repo_path))
                        })
                        break
    
    elif language == 'python':
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            
            for pattern in api_patterns:
                matches = re.findall(pattern, line)
                for match in matches:
                    api_name = match
                    if api_name and not api_name.startswith('_'):
                        apis.append({
                            'name': api_name,
                            'source': source_file.name,
                            'category': category,
                            'full_path': str(source_file.relative_to(repo_path))
                        })
                        break
    
    elif language == 'javascript':
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('/*'):
                continue
            
            for pattern in api_patterns:
                matches = re.findall(pattern, line)
                for match in matches:
                    api_name = match
                    if api_name:
                        apis.append({
                            'name': api_name,
                            'source': source_file.name,
                            'category': category,
                            'full_path': str(source_file.relative_to(repo_path))
                        })
                        break
    
    elif language == 'java':
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('/*'):
                continue
            
            for pattern in api_patterns:
                matches = re.findall(pattern, line)
                for match in matches:
                    api_name = match[1] if isinstance(match, tuple) else match
                    if api_name and not api_name.startswith('_'):
                        apis.append({
                            'name': api_name,
                            'source': source_file.name,
                            'category': category,
                            'full_path': str(source_file.relative_to(repo_path))
                        })
                        break
    
    elif language == 'go':
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('//'):
                continue
            
            for pattern in api_patterns:
                matches = re.findall(pattern, line)
                for match in matches:
                    api_name = match
                    if api_name and api_name[0].isupper():
                        apis.append({
                            'name': api_name,
                            'source': source_file.name,
                            'category': category,
                            'full_path': str(source_file.relative_to(repo_path))
                        })
                        break
    
    return apis

def check_test_coverage(api_name, test_files):
    matching_tests = []
    
    for test_file in test_files:
        try:
            content = test_file.read_text(encoding='utf-8', errors='ignore')
            if api_name in content:
                matching_tests.append(test_file.name)
        except Exception:
            pass
    
    if matching_tests:
        return 'Has test', ', '.join(list(set(matching_tests)))
    else:
        return 'No test', '-'

def generate_text_report(apis, repo_path, output_path):
    categories = {}
    for api in apis:
        cat = api['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(api)
    
    total_apis = len(apis)
    tested_apis = sum(1 for api in apis if api['test_status'] == 'Has test')
    coverage = (tested_apis / total_apis * 100) if total_apis > 0 else 0
    
    report = []
    report.append('=' * 100)
    report.append(f"{repo_path.name} Test Coverage Report")
    report.append('=' * 100)
    report.append("")
    report.append(f"Repository: {repo_path}")
    report.append(f"Total APIs: {total_apis}")
    report.append(f"Tested APIs: {tested_apis}")
    report.append(f"Coverage: {coverage:.1f}%")
    report.append("")
    report.append('=' * 100)
    report.append("")
    report.append("【API Coverage by Category】")
    report.append("-" * 100)
    report.append(f"{'Category':<20} {'APIs':<10} {'Tested':<10} {'Coverage'}")
    report.append("-" * 100)
    
    for cat, cat_apis in sorted(categories.items()):
        cat_tested = sum(1 for api in cat_apis if api['test_status'] == 'Has test')
        cat_coverage = (cat_tested / len(cat_apis) * 100) if cat_apis else 0
        report.append(f"{cat:<20} {len(cat_apis):<10} {cat_tested:<10} {cat_coverage:.1f}%")
    
    report.append("")
    report.append('=' * 100)
    report.append("")
    report.append("【API Details】")
    report.append("-" * 100)
    report.append(f"{'Category':<20} {'API Name':<40} {'Source File':<30} {'Test Status':<15} {'Test File'}")
    report.append("-" * 100)
    
    for api in apis:
        report.append(f"{api['category']:<20} {api['name']:<40} {api['source']:<30} {api['test_status']:<15} {api['test_file']}")
    
    report.append("")
    report.append('=' * 100)
    report.append("")
    report.append("【Missing Tests】")
    report.append("-" * 100)
    
    missing_tests = [api for api in apis if api['test_status'] == 'No test']
    if missing_tests:
        report.append(f"Total APIs without tests: {len(missing_tests)}")
        report.append("")
        for api in missing_tests:
            report.append(f"  - {api['name']} ({api['category']})")
    else:
        report.append("All APIs have tests!")
    
    report.append("")
    report.append("=" * 100)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))
    
    return output_path

def generate_excel_report(apis, repo_path, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not available, skipping Excel report generation")
        return None
    
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    
    ws_apis = wb.create_sheet("APIs by Category")
    ws_apis.append(["Category", "API Name", "Source File", "Test Status", "Test File"])
    for api in apis:
        ws_apis.append([api['category'], api['name'], api['source'], api['test_status'], api['test_file']])
    
    ws_stats = wb.create_sheet("Statistics")
    ws_stats.append(["Category", "Total APIs", "Tested APIs", "Coverage %"])
    
    categories = {}
    for api in apis:
        cat = api['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(api)
    
    for cat, cat_apis in sorted(categories.items()):
        cat_tested = sum(1 for api in cat_apis if api['test_status'] == 'Has test')
        cat_coverage = (cat_tested / len(cat_ars) * 100) if cat_apis else 0
        ws_stats.append([cat, len(cat_apis), cat_tested, f"{cat_coverage:.1f}%"])
    
    total_apis = len(apis)
    tested_apis = sum(1 for api in apis if api['test_status'] == 'Has test')
    coverage = (tested_apis / total_apis * 100) if total_apis > 0 else 0
    
    ws_stats.append([])
    ws_stats.append(["Overall", total_apis, tested_apis, f"{coverage:.1f}%"])
    
    ws_missing = wb.create_sheet("Missing Tests")
    ws_missing.append(["API Name", "Category", "Source File"])
    
    missing_tests = [api for api in apis if api['test_status'] == 'No test']
    for api in missing_tests:
        ws_missing.append([api['name'], api['category'], api['source']])
    
    wb.save(output_path)
    return output_path

def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_coverage.py <repo_path> [output_dir]")
        print("Example: python analyze_coverage.py /path/to/repo /path/to/output")
        sys.exit(1)
    
    repo_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else repo_path
    
    print(f"Analyzing repository: {repo_path}")
    
    language = detect_language(repo_path)
    print(f"Detected language: {language}")
    
    source_files = find_source_files(repo_path, language)
    print(f"Found {len(source_files)} source files")
    
    test_files = find_test_files(repo_path, language)
    print(f"Found {len(test_files)} test files")
    
    all_apis = []
    for source_file in source_files:
        apis = extract_apis_from_source(source_file, repo_path, language)
        all_apis.extend(apis)
    
    print(f"Extracted {len(all_apis)} APIs")
    
    for api in all_apis:
        test_status, test_file = check_test_coverage(api['name'], test_files)
        api['test_status'] = test_status
        api['test_file'] = test_file
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    text_report_path = output_dir / f"{repo_path.name}_test_coverage_report.txt"
    excel_report_path = output_dir / f"{repo_path.name}_api_coverage.xlsx"
    
    text_path = generate_text_report(all_apis, repo_path, str(text_report_path))
    excel_path = generate_excel_report(all_apis, repo_path, str(excel_report_path))
    
    total_apis = len(all_apis)
    tested_apis = sum(1 for api in all_apis if api['['test_status'] == 'Has test')
    coverage = (tested_apis / total_apis * 100) if total_apis > 0 else 0
    
    print(f"\nSummary:")
    print(f"  Total APIs: {total_apis}")
    print(f"  Tested APIs: {tested_apis}")
    print(f"  Coverage: {coverage:.1f}%")
    print(f"\nGenerated reports:")
    print(f"  Text report: {text_path}")
    print(f"  Excel report: {excel_path}")

if __name__ == "__main__":
    main()
